# work with the excel module and operation

import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]
product_per_supplier = {}
total_value_per_supplier = {}
product_under_10 = {}
# how many product we have in the sheet number of the product per supplier with looping

for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_num = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 5) # no value cause we need whole object
    if supplier_name in product_per_supplier:
        product_per_supplier[supplier_name] = product_per_supplier[supplier_name] + 1
    else:

        product_per_supplier[supplier_name] = 1  # it will add the supplier to dict

    # calculation total value of inventory per supplier
    if supplier_name in total_value_per_supplier:
        total_value_per_supplier[supplier_name] = total_value_per_supplier[supplier_name] + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price
    # product under inventory 10 will contain prod number and inventory value

    if inventory < 10:
        product_under_10[product_num] = inventory

    # add total price inventory * price
    inventory_price.value = inventory * price


print(product_per_supplier)
print(total_value_per_supplier)
print(product_under_10)

inv_file.save("new_inventory.xlsx")













